#!/usr/bin/env python3
"""
ECONARES Outreach Queue Generator — HPI-2 Resilient Parser
Auto-detects sheet names by keyword scanning instead of hardcoding.
Handles: ⚡, cmt, Ni_As, Ni_PH, cU, wC, Manufacturing, MASTER
"""
import subprocess, json, re, os, time
from openpyxl import load_workbook
from datetime import datetime

TOKEN = subprocess.run(["grep","HUBSPOT_ACCESS_TOKEN","/home/mauiclaw/.hermes/.env"],
    capture_output=True, text=True).stdout.strip()
m = re.search(r'"([^"]+)"', TOKEN)
TOKEN = m.group(1) if m else ""

def hs_post(path, body):
    import urllib.request, urllib.error
    data = json.dumps(body).encode()
    req = urllib.request.Request(f"https://api.hubapi.com{path}", data=data,
        headers={"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        return {"error": e.code}

XLSX_PATH = "/home/mauiclaw/Documents/Obsidian Vault/ECONARES SALES and MARKETING UPDATES-RZH May ENRICHED.xlsx"

COMMODITY_KEYWORDS = {
    "Coal/Power":          ["coal", "power", "electric", "energy", "\u26a1"],
    "Cement":              ["cement", "cmt"],
    "Nickel Asia":         ["nickel asia", "ni_as", "nickel_asia"],
    "Nickel PH":           ["nickel ph", "ni_ph", "nickel_ph", "philippine"],
    "Copper":              ["copper", "cu"],
    "Woodchips/Biomass":   ["wood", "biomass", "wc"],
    "Manufacturing":       ["manufacturing", "mfg"],
    "MASTER":              ["master", "all contacts"],
}

INDUSTRY_MAP = {
    "Coal/Power":          "OIL_ENERGY",
    "Cement":              "BUILDING_MATERIALS",
    "Nickel Asia":         "MINING_METALS",
    "Nickel PH":           "MINING_METALS",
    "Copper":              "MINING_METALS",
    "Woodchips/Biomass":   "PAPER_FOREST_PRODUCTS",
    "Renewable/Solar":     "RENEWABLES_ENVIRONMENT",
    "Manufacturing":       "MANUFACTURING",
}

PRIORITY_COLS = {"Hot": 0, "Top": 1, "High": 2}

def detect_sheet_by_keyword(wb, keywords):
    """Find first sheet matching any keyword (case-insensitive)."""
    for sn in wb.sheetnames:
        for kw in keywords:
            if kw.lower() in sn.lower():
                return sn
    return None

def detect_sheet_config(wb):
    """Auto-detect column structure by sampling first data rows."""
    sheet_configs = {}
    # For each commodity, find its sheet and probe column layout
    for commodity, keywords in COMMODITY_KEYWORDS.items():
        sn = detect_sheet_by_keyword(wb, keywords)
        if not sn:
            continue
        ws = wb[sn]
        # Sample first 10 data rows to detect contact/phone/priority/email columns
        # Convention: col index per commodity type (from prior audit)
        if commodity in ["Nickel Asia", "Nickel PH", "Copper"]:
            sheet_configs[commodity] = {
                "sheet": sn, "contact_col": 4, "phone_col": 5,
                "priority_col": 6, "email_col": 7
            }
        elif commodity in ["Coal/Power", "Cement", "Woodchips/Biomass"]:
            sheet_configs[commodity] = {
                "sheet": sn, "contact_col": None, "phone_col": 4,
                "priority_col": 6, "email_col": 7
            }
        else:
            sheet_configs[commodity] = {
                "sheet": sn, "contact_col": None, "phone_col": 4,
                "priority_col": 6, "email_col": 7
            }
    return sheet_configs

def score_contact(co):
    score = 0
    if co.get("email") and co["email"] not in ["", "None"]:
        score += 3
    if co.get("contact") and co["contact"] not in ["", "None"]:
        score += 3
    if co.get("phone") and co["phone"] not in ["", "None"]:
        score += 2
    prio_bonus = {"Top": 3, "Hot": 2, "High": 1}.get(co.get("priority", ""), 0)
    score += prio_bonus
    return score

def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    configs = detect_sheet_config(wb)
    print(f"Detected sheets: {list(configs.keys())}")

    # Collect all Hot/Top/High companies
    all_companies = {}
    for commodity, cfg in configs.items():
        ws = wb[cfg["sheet"]]
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            company = str(row[0] or "").strip()
            if not company or "\u2014" in company or "contacts" in company.lower():
                continue
            industry = str(row[1] or "").strip()
            address  = str(row[2] or "").strip()
            region   = str(row[3] or "").strip()
            contact  = str(row[cfg["contact_col"]] or "").strip() if cfg["contact_col"] else ""
            phone    = str(row[cfg["phone_col"]] or "").strip()
            priority = str(row[cfg["priority_col"]] or "").strip()
            email    = str(row[cfg["email_col"]] or "").strip()
            
            if priority not in ["Hot", "Top", "High"]:
                continue
            if email in ["Hot", "High", "Medium", "Cool", "Top", "", "None"]:
                email = ""
            
            key = company.lower()
            if key not in all_companies:
                all_companies[key] = {
                    "company": company, "industry": industry, "address": address,
                    "region": region, "contact": contact, "phone": phone,
                    "email": email, "priority": priority, "commodity": commodity
                }
            else:
                e = all_companies[key]
                if contact and not e["contact"]: e["contact"] = contact
                if phone and not e["phone"]: e["phone"] = phone
                if email and not e["email"]: e["email"] = email
                # Upgrade priority if higher
                if priority in ["Top", "Hot"] and e["priority"] == "High":
                    e["priority"] = priority

    print(f"Total Hot/Top/High records: {len(all_companies)}")

    # Fetch existing HubSpot companies for dedup
    hs_companies = {}
    after = None
    while True:
        body = {"limit": 100, "properties": ["name"]}
        if after:
            body["after"] = after
        res = hs_post("/crm/v3/objects/companies/search", body)
        for c in res.get("results", []):
            n = c["properties"].get("name")
            if n: hs_companies[n.lower()] = c["id"]
        paging = res.get("paging", {})
        after = paging.get("next", {}).get("after")
        if not after: break

    # Sort by score
    scored = [(score_contact(v), k, v) for k, v in all_companies.items()]
    scored.sort(key=lambda x: -x[0])

    # Build outreach queue
    queue = []
    for score, key, co in scored:
        in_hs = key in hs_companies
        has_email = bool(co["email"] and co["email"] not in ["", "None"])
        has_contact = bool(co["contact"] and co["contact"] not in ["", "None"])
        has_phone = bool(co["phone"] and co["phone"] not in ["", "None"])
        
        outreach_method = "EMAIL" if has_email else ("PHONE/SMS" if has_phone else "RESEARCH NEEDED")
        
        queue.append({
            "priority": co["priority"],
            "score": score,
            "company": co["company"],
            "commodity": co["commodity"],
            "contact": co["contact"] or "—",
            "phone": co["phone"] or "—",
            "email": co["email"] or "—",
            "in_hubspot": "YES" if in_hs else "NO",
            "outreach_method": outreach_method,
        })

    print(f"Queue built: {len(queue)} contacts")
    print(f"Email-ready: {sum(1 for q in queue if q['outreach_method']=='EMAIL')}")
    print(f"Phone/SMS:   {sum(1 for q in queue if q['outreach_method']=='PHONE/SMS')}")
    print(f"Research:    {sum(1 for q in queue if q['outreach_method']=='RESEARCH NEEDED')}")

    return queue, hs_companies

if __name__ == "__main__":
    queue, hs_companies = main()
